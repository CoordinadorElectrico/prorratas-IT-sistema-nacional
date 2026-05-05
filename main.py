# Librerías
import pandas as pd
from openpyxl import load_workbook

# Carpetas y archivos de datos
# Carpeta donde se guardan las salidas de la liquidación
ruta_salidas='U:\\Pagos por Transmisión Nacional\\2025\\04_Liquidación\\2503\\02 Salidas\\Preliminar\\'

# Carpeta donde se guarda la planilla ENT
ruta_planilla_ENT='U:\\Pagos por Transmisión Nacional\\2025\\04_Liquidación\\2503\\01 Entradas\\Preliminar\\'

# Carpeta donde se guarda planilla de prorratas
ruta_prorratas='U:\\Pagos por Transmisión Nacional\\2025\\04_Liquidación\\2503\\03 Prorratas Facturacion\\Preliminar\\'

# Mes de cálculo
mes='Mar'         # <<<<<<< ---------- CAMBIAR MES CÁLCULO ----------

# Nombres de archivos
# 
PjeClienLin='PjeClienLin2025.csv'
CMes='CMes2025.csv'
PagoIny='PagoIny2025.csv'
PlanillaENT='Ent2025.xlsx'
PlanillaProrratas='01. ProrratasPago_Mar25_pre_prueba.xlsx' # <<<<<<< ---------- NOMBRE PLANILLA CON TABLAS ----------

# Lectura CMes.csv
CMes=pd.read_csv(ruta_salidas + CMes, encoding='latin')
CMes=CMes.loc[CMes.Mes == mes]
CMes[['Cliente_Nuevo','Suministrador','Barra']]=CMes['Cliente'].str.split('#', expand=True)
CMes.reset_index(inplace=True)
consumo_x_barra=pd.DataFrame(CMes.groupby(['Barra'])['Consumo[MWh]'].sum())
consumo_x_barra.reset_index(inplace=True)

# Lectura PagoIny.csv
PagoIny=pd.read_csv(ruta_salidas + PagoIny, encoding='latin')
PagoIny=PagoIny.loc[PagoIny.Mes == mes]
PagoIny['PagoFinal'] = PagoIny.apply(lambda row: row['Pago de Peaje [$]'] * row['FAER']*row['CET']*(1-row['Factor de Exención']) if row['FAER'] <= row['Tabla 1'] else row['Pago de Peaje [$]'] * row['Tabla 1']*row['CET']*(1-row['Factor de Exención']), axis=1)
PagoIny_agregado = PagoIny.groupby(['Tramo', 'Transmisor', 'Empresa Generación'])[['Pago de Peaje [$]', 'PagoFinal']].sum().reset_index()
PagoIny_agregado.rename(columns={'Empresa Generación': 'Suministrador'}, inplace=True)

# Lectura PjeClienLin.csv
PjeClienLin=pd.read_csv(ruta_salidas + PjeClienLin, usecols=['Tramo','Transmisor','Cliente','Suministrador', 'Barra', 'Mes', 'PagoxLinea[$]'], encoding='latin-1')
PjeClienLin=PjeClienLin.loc[PjeClienLin.Mes == mes]
PagoxTramoxBarra = PjeClienLin.groupby(['Tramo', 'Transmisor', 'Barra'])['PagoxLinea[$]'].sum().reset_index()
PjeClienLin=PjeClienLin.merge(PagoxTramoxBarra, on=['Tramo', 'Transmisor', 'Barra'], how='left', suffixes=(None,'_y'))
PjeClienLin=PjeClienLin.merge(consumo_x_barra, on='Barra', how='left', suffixes=(None,'_z'))
PjeClienLin['PUxBarra'] = PjeClienLin.apply(lambda row: row['PagoxLinea[$]_y'] / row['Consumo[MWh]'] if row['Consumo[MWh]'] != 0 else 0, axis=1)
PjeClienLin=PjeClienLin.merge(CMes, left_on=['Cliente', 'Suministrador', 'Barra','Mes'], right_on=['Cliente_Nuevo', 'Suministrador', 'Barra','Mes'],how='left', suffixes=(None,'_w'))
PjeClienLin['PagoPUxTramo'] = PjeClienLin['PUxBarra']*PjeClienLin['Consumo[MWh]_w']
PjeClienLin['PagoCI'] = PjeClienLin.apply(lambda row: row['PagoPUxTramo'] if row['Suministrador'] != 'SUMINISTRADOR_NO_IND' else 0, axis=1)
PjeClienLin['PagoCNI'] = PjeClienLin.apply(lambda row: row['PagoPUxTramo'] if row['Suministrador'] == 'SUMINISTRADOR_NO_IND' else 0, axis=1)
PjeClienLin_agregado = PjeClienLin.groupby(['Tramo', 'Transmisor', 'Suministrador'])['PagoPUxTramo'].sum().reset_index()
PjeClienLin_agregado.rename(columns={'PagoPUxTramo': 'PeajeRetiro'}, inplace=True)

# PagoxTramoTotal
PagoxTramo = pd.concat([PjeClienLin_agregado, PagoIny_agregado], ignore_index=True)
PagoxTramo.fillna(0, inplace=True)
PagoxTramo['PIny_ini+PRet']=PagoxTramo['Pago de Peaje [$]'] + PagoxTramo['PeajeRetiro']
PagoxTramoTotal = PagoxTramo.groupby(['Tramo', 'Transmisor'])['PIny_ini+PRet'].sum().reset_index()

# Tablas Peajes de Inyección, Retiro y Resumen
Prorratas_Resumen=PagoxTramo.merge(PagoxTramoTotal, on=['Tramo', 'Transmisor'], how='left', suffixes=(None,'_u'))
Prorratas_Resumen['ProrrataPago'] = Prorratas_Resumen.apply(lambda row: row['PIny_ini+PRet']/row['PIny_ini+PRet_u'] if row['PIny_ini+PRet_u'] != 0 else 0, axis=1)
Prorratas_Resumen['ProrrataPagoPIny'] = Prorratas_Resumen.apply(lambda row: row['PagoFinal']/row['PIny_ini+PRet_u'] if row['PIny_ini+PRet_u'] != 0 else 0, axis=1)
Prorratas_Resumen['ProrrataPagoPRetCI'] = Prorratas_Resumen.apply(lambda row: row['PeajeRetiro']/row['PIny_ini+PRet_u'] if (row['PIny_ini+PRet_u'] != 0) and (row['Suministrador'] != 'SUMINISTRADOR_NO_IND') else 0, axis=1)
Prorratas_Resumen['ProrrataPagoPRetCNI'] = Prorratas_Resumen.apply(lambda row: row['PeajeRetiro']/row['PIny_ini+PRet_u'] if (row['PIny_ini+PRet_u'] != 0) and (row['Suministrador'] == 'SUMINISTRADOR_NO_IND') else 0, axis=1)
Prorratas_Resumen['ProrrataPagoPInyIni'] = Prorratas_Resumen.apply(lambda row: row['Pago de Peaje [$]']/row['PIny_ini+PRet_u'] if row['PIny_ini+PRet_u'] != 0 else 0, axis=1)
Prorratas_Resumen['ProrrataExenc']=Prorratas_Resumen['ProrrataPagoPInyIni'] - Prorratas_Resumen['ProrrataPagoPIny']
Prorratas_Resumen['Tramo#Transmisor']=Prorratas_Resumen['Tramo'] + '#' + Prorratas_Resumen['Transmisor']
Tabla_PIny_PRet=Prorratas_Resumen.loc[Prorratas_Resumen.Suministrador != 'SUMINISTRADOR_NO_IND']

# Tabla PINY
tabla_PINY = pd.pivot_table(Tabla_PIny_PRet, values='ProrrataPagoPIny', index=['Suministrador', 'Tramo#Transmisor'], aggfunc='sum').reset_index()
tabla_PINY=tabla_PINY.loc[tabla_PINY.ProrrataPagoPIny != 0]

# Tabla PRET
tabla_PRET = pd.pivot_table(Tabla_PIny_PRet, values='ProrrataPagoPRetCI', index=['Suministrador', 'Tramo#Transmisor'], aggfunc='sum').reset_index()
tabla_PRET=tabla_PRET.loc[tabla_PRET.ProrrataPagoPRetCI != 0]

# Tabla Resumen
TablaResumen=Prorratas_Resumen.groupby('Tramo#Transmisor')[['ProrrataPagoPIny','ProrrataExenc','ProrrataPagoPRetCNI','ProrrataPagoPRetCI']].sum().reset_index()

# Lectuta de planilla ENT
wb=load_workbook(filename=ruta_planilla_ENT + PlanillaENT, data_only=True)
nombre_rango='barcli'
rango = wb.defined_names[nombre_rango]
destinos = list(rango.destinations)
hoja_nombre, celdas = destinos[0]
hoja = wb[hoja_nombre]
celda_inicio, celda_fin = celdas.split(':')
datos=[]
for fila in hoja[celda_inicio:celda_fin]:
    datos.append([celda.value for celda in fila])
Clientes_CI = pd.DataFrame(datos)
Clientes_CI.columns = ['indice', 'tripleta', 'Cliente', 'Suministrador', 'Barra', 'TipoCliente']
Clientes_CI=Clientes_CI.loc[Clientes_CI.TipoCliente == 1]
columnas_deseadas = ['Cliente', 'Suministrador', 'Barra']

# Tabla Clientes Individualizados
Clientes_CI=Clientes_CI[columnas_deseadas]

# Crea libro Excel
with pd.ExcelWriter(ruta_prorratas + PlanillaProrratas, engine='openpyxl') as writer:
    TablaResumen.to_excel(writer, sheet_name='Resumen', index=False)
    Clientes_CI.to_excel(writer, sheet_name='Clientes Ind', index=False)
    tabla_PINY.to_excel(writer, sheet_name='PINY', index=False)
    tabla_PRET.to_excel(writer, sheet_name='PRET', index=False)

# Ajusta columnas
wb = load_workbook(ruta_prorratas + PlanillaProrratas)
for hoja in wb.sheetnames:
    ws = wb[hoja]
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[col_letter].width = adjusted_width

wb.save(ruta_prorratas + PlanillaProrratas)
